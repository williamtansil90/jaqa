from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.models import (
    Expectation,
    Step,
    EXPECT_EXPORT_FIELDS,
    STEP_EXPORT_FIELDS,
    TC_FIELD_LABELS,
    TC_FILE_FIELDS,
    TC_LIST_COLUMNS,
    TestCase,
    TestSuite,
    _as_enabled,
)

HEADERS = [label for _, label in TC_LIST_COLUMNS]
TC_FILE_HEADERS = [TC_FIELD_LABELS[key] for key in TC_FILE_FIELDS]
STEP_FILE_HEADERS = [label for _, label in STEP_EXPORT_FIELDS]
EXPECT_FILE_HEADERS = [label for _, label in EXPECT_EXPORT_FIELDS]


def _expectation_summary(case: TestCase) -> str:
    if not case.expectations:
        return ""
    return "\n".join(f"• {item.summary()}" for item in case.expectations)


def _rows(suite: TestSuite) -> list[list[str]]:
    rows: list[list[str]] = []
    for case in suite.test_cases:
        rows.append(
            [
                "ENABLE" if case.enabled else "DISABLE",
                case.no_tc,
                case.deskripsi,
                case.aplikasi,
                case.url,
                case.username,
                case.password,
                case.expected_result,
                _expectation_summary(case),
                case.status or "BELUM DIUJI",
                case.notes,
            ]
        )
    return rows


def export_excel(suite: TestSuite, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.title = "Hasil SIT"

    title_fill = PatternFill("solid", fgColor="0B3A4A")
    header_fill = PatternFill("solid", fgColor="115E67")
    ok_fill = PatternFill("solid", fgColor="C6F4D6")
    nok_fill = PatternFill("solid", fgColor="FECACA")
    pending_fill = PatternFill("solid", fgColor="E5E7EB")
    alt_fill = PatternFill("solid", fgColor="F0FDFA")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    white = Font(color="FFFFFF", bold=True, name="Calibri")
    wrap = Alignment(wrap_text=True, vertical="center")

    sheet.merge_cells("A1:K1")
    sheet["A1"] = "JAQA — Jalin Automate QA  |  Laporan Hasil SIT"
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:K2")
    ok_n = sum(1 for tc in suite.test_cases if tc.status == "OK")
    nok_n = sum(1 for tc in suite.test_cases if tc.status == "NOK")
    pending_n = len(suite.test_cases) - ok_n - nok_n
    sheet["A2"] = (
        f"Suite: {suite.name}   •   Diekspor: {datetime.now().strftime('%d %b %Y %H:%M')}   •   "
        f"Total {len(suite.test_cases)} TC   •   OK {ok_n}   •   NOK {nok_n}   •   Belum diuji {pending_n}"
    )
    sheet["A2"].font = Font(color="134E4A", name="Calibri", size=10)
    sheet["A2"].alignment = Alignment(vertical="center", indent=1)
    sheet.row_dimensions[2].height = 20

    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(3, col, header)
        cell.fill = header_fill
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    sheet.row_dimensions[3].height = 22
    sheet.auto_filter.ref = f"A3:K{3 + max(len(suite.test_cases), 1)}"
    sheet.freeze_panes = "A4"

    for row_idx, values in enumerate(_rows(suite), start=4):
        status = values[9]
        fill = ok_fill if status == "OK" else nok_fill if status == "NOK" else pending_fill
        sheet.row_dimensions[row_idx].height = 36
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_idx, col, value)
            cell.alignment = wrap
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if col == 10:
                cell.fill = fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="14532D" if status == "OK" else "7F1D1D" if status == "NOK" else "334155")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    widths = [12, 14, 36, 20, 34, 16, 16, 28, 42, 14, 46]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    book.save(target)
    return target


def export_pdf(suite: TestSuite, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "JAQATitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor("#0B3A4A"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "JAQAMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#115E67"),
        spaceAfter=10,
    )
    cell = ParagraphStyle(
        "JAQACell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#0F172A"),
    )
    cell_center = ParagraphStyle("JAQACellCenter", parent=cell, alignment=TA_CENTER, fontName="Helvetica-Bold")

    ok_n = sum(1 for tc in suite.test_cases if tc.status == "OK")
    nok_n = sum(1 for tc in suite.test_cases if tc.status == "NOK")
    pending_n = len(suite.test_cases) - ok_n - nok_n

    def p(text: str, style=cell) -> Paragraph:
        safe = (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
        return Paragraph(safe, style)

    header_style = ParagraphStyle(
        "JAQAHeader",
        parent=cell_center,
        textColor=colors.white,
        fontSize=8,
    )
    header = [p(h, header_style) for h in HEADERS]
    data = [header]
    status_colors: list[str] = []
    for values in _rows(suite):
        status = values[9]
        status_colors.append(status)
        styled = [p(v) for v in values]
        styled[9] = p(status, cell_center)
        data.append(styled)

    doc = SimpleDocTemplate(
        str(target),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="JAQA Laporan SIT",
        author="JAQA — Jalin Automate QA",
    )
    table = Table(
        data,
        colWidths=[16 * mm, 22 * mm, 38 * mm, 28 * mm, 40 * mm, 24 * mm, 24 * mm, 32 * mm, 48 * mm, 20 * mm, 48 * mm],
        repeatRows=1,
    )
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#115E67")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (9, 0), (9, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#94A3B8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for idx, status in enumerate(status_colors, start=1):
        if status == "OK":
            commands.append(("BACKGROUND", (9, idx), (9, idx), colors.HexColor("#86EFAC")))
        elif status == "NOK":
            commands.append(("BACKGROUND", (9, idx), (9, idx), colors.HexColor("#FCA5A5")))
        else:
            commands.append(("BACKGROUND", (9, idx), (9, idx), colors.HexColor("#E2E8F0")))
        if idx % 2 == 0:
            commands.append(("BACKGROUND", (0, idx), (8, idx), colors.HexColor("#F0FDFA")))
            commands.append(("BACKGROUND", (10, idx), (10, idx), colors.HexColor("#F0FDFA")))
    table.setStyle(TableStyle(commands))

    story = [
        Paragraph("JAQA — Jalin Automate QA", title_style),
        Paragraph(
            f"Laporan Hasil SIT &nbsp;&nbsp;|&nbsp;&nbsp; {suite.name} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"{datetime.now().strftime('%d %b %Y %H:%M')} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Total {len(suite.test_cases)} TC &nbsp;&nbsp;•&nbsp;&nbsp; OK {ok_n} &nbsp;&nbsp;•&nbsp;&nbsp; "
            f"NOK {nok_n} &nbsp;&nbsp;•&nbsp;&nbsp; Belum diuji {pending_n}",
            meta_style,
        ),
        table,
        Spacer(1, 8),
        Paragraph(
            "Status OK (hijau) = semua expected result sesuai. Status NOK (merah) = ada expected yang gagal; lihat kolom Notes.",
            meta_style,
        ),
    ]
    doc.build(story)
    return target


def _export_values(fields: tuple[tuple[str, str], ...], row: dict[str, object]) -> list[object]:
    return [row.get(key, "") for key, _label in fields]


def _step_export_row(case_no: str, order: int, step: Step) -> dict[str, object]:
    checked = "" if step.checked is None else ("TRUE" if step.checked else "FALSE")
    return {
        "no_tc": case_no,
        "urutan": order,
        "enabled": "ENABLE" if step.enabled else "DISABLE",
        "type": step.type,
        "label": step.label,
        "selector": step.selector,
        "value": step.value,
        "url": step.url,
        "key": step.key,
        "checked": checked,
        "delay_ms": step.delay_ms,
        "tag": step.tag,
        "id": step.id,
    }


def _expectation_export_row(case_no: str, order: int, item: Expectation) -> dict[str, object]:
    return {
        "no_tc": case_no,
        "urutan": order,
        "enabled": "ENABLE" if item.enabled else "DISABLE",
        "label": item.label,
        "selector": item.selector,
        "kind": item.kind,
        "match": item.match,
        "expected_value": item.expected_value,
        "attribute": item.attribute,
        "after_step": item.after_step,
        "tag": item.tag,
        "sample_text": item.sample_text,
        "id": item.id,
    }


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _header_key(value: str) -> str | None:
    text = (value or "").strip()
    if text == "#":
        return "urutan"
    aliases = {
        "id": "id",
        "active": "enabled",
        "aktif": "enabled",
        "enabled": "enabled",
        "enable": "enabled",
        "notc": "no_tc",
        "notestcase": "no_tc",
        "tc": "no_tc",
        "deskripsi": "deskripsi",
        "description": "deskripsi",
        "aplikasi": "aplikasi",
        "application": "aplikasi",
        "app": "aplikasi",
        "url": "url",
        "username": "username",
        "user": "username",
        "password": "password",
        "pass": "password",
        "ekspetasi": "expected_result",
        "expectation": "expected_result",
        "expectedresult": "expected_result",
        "expected": "expected_result",
        "notes": "notes",
        "note": "notes",
        "catatan": "notes",
        "status": "status",
        "urutan": "urutan",
        "order": "urutan",
        "step": "label",
        "keterangan": "label",
        "type": "type",
        "tipe": "type",
        "label": "label",
        "selector": "selector",
        "value": "value",
        "nilai": "value",
        "key": "key",
        "checked": "checked",
        "delayms": "delay_ms",
        "delay": "delay_ms",
        "jeda": "delay_ms",
        "tag": "tag",
        "kind": "kind",
        "jenis": "kind",
        "match": "match",
        "comparison": "match",
        "banding": "match",
        "expectedvalue": "expected_value",
        "attribute": "attribute",
        "afterstep": "after_step",
        "checkafter": "after_step",
        "perikasaetelah": "after_step",
        "periksasetelah": "after_step",
        "sample": "sample_text",
    }
    return aliases.get(_norm_header(value))


def _style_header_row(sheet, headers: list[str], fill: PatternFill, font: Font, border: Border) -> None:
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.row_dimensions[1].height = 22
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def export_tc_excel(suite: TestSuite, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    header_fill = PatternFill("solid", fgColor="115E67")
    white = Font(color="FFFFFF", bold=True, name="Calibri")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    wrap = Alignment(wrap_text=True, vertical="center")

    cases = book.active
    cases.title = "Test Cases"
    _style_header_row(cases, TC_FILE_HEADERS + ["ID"], header_fill, white, thin)
    for row_idx, case in enumerate(suite.test_cases, start=2):
        values = [
            "ENABLE" if case.enabled else "DISABLE",
            case.no_tc,
            case.deskripsi,
            case.aplikasi,
            case.url,
            case.username,
            case.password,
            case.expected_result,
            case.notes,
            case.id,
        ]
        for col, value in enumerate(values, start=1):
            cell = cases.cell(row_idx, col, value)
            cell.alignment = wrap
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
        cases.row_dimensions[row_idx].height = 28
    for idx, width in enumerate([12, 14, 36, 20, 36, 16, 16, 28, 46, 16], start=1):
        cases.column_dimensions[get_column_letter(idx)].width = width

    steps = book.create_sheet("Steps")
    _style_header_row(steps, STEP_FILE_HEADERS, header_fill, white, thin)
    step_row = 2
    for case in suite.test_cases:
        for order, step in enumerate(case.steps, start=1):
            values = _export_values(STEP_EXPORT_FIELDS, _step_export_row(case.no_tc, order, step))
            for col, value in enumerate(values, start=1):
                cell = steps.cell(step_row, col, value)
                cell.alignment = wrap
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            step_row += 1
    for idx, width in enumerate([14, 8, 12, 12, 24, 28, 20, 28, 10, 12, 12, 12, 16], start=1):
        steps.column_dimensions[get_column_letter(idx)].width = width

    expects = book.create_sheet("Expectations")
    _style_header_row(expects, EXPECT_FILE_HEADERS, header_fill, white, thin)
    exp_row = 2
    for case in suite.test_cases:
        for order, item in enumerate(case.expectations, start=1):
            values = _export_values(EXPECT_EXPORT_FIELDS, _expectation_export_row(case.no_tc, order, item))
            for col, value in enumerate(values, start=1):
                cell = expects.cell(exp_row, col, value)
                cell.alignment = wrap
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            exp_row += 1
    for idx, width in enumerate([14, 8, 12, 22, 28, 12, 14, 28, 14, 14, 12, 20, 16], start=1):
        expects.column_dimensions[get_column_letter(idx)].width = width

    book.save(target)
    return target


def _sheet_maps(sheet) -> list[dict[str, str]]:
    header_row = None
    mapping: dict[int, str] = {}
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 8), values_only=False):
        keys = [_header_key("" if cell.value is None else str(cell.value)) for cell in row]
        if "no_tc" in keys or "deskripsi" in keys or ("urutan" in keys and ("type" in keys or "kind" in keys)):
            header_row = row[0].row
            mapping = {cell.column: key for cell, key in zip(row, keys) if key}
            break
    if not header_row or not mapping:
        return []
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
        item: dict[str, str] = {}
        empty = True
        for cell in row:
            key = mapping.get(cell.column)
            if not key:
                continue
            text = "" if cell.value is None else str(cell.value).strip()
            item[key] = text
            if text:
                empty = False
        if not empty:
            rows.append(item)
    return rows


def import_tc_excel(path: str | Path) -> TestSuite:
    book = load_workbook(filename=str(path), data_only=True)
    case_sheet = None
    for name in ("Test Cases", "TestCases", "TC", "Sheet1"):
        if name in book.sheetnames:
            case_sheet = book[name]
            break
    if case_sheet is None:
        case_sheet = book.worksheets[0]

    cases: list[TestCase] = []
    by_no: dict[str, TestCase] = {}
    for item in _sheet_maps(case_sheet):
        no_tc = item.get("no_tc") or ""
        case = TestCase(
            id=item.get("id") or "",
            no_tc=no_tc,
            deskripsi=item.get("deskripsi", ""),
            aplikasi=item.get("aplikasi", ""),
            url=item.get("url", ""),
            username=item.get("username", ""),
            password=item.get("password", ""),
            expected_result=item.get("expected_result", ""),
            notes=item.get("notes", ""),
            enabled=_as_enabled(item.get("enabled", True)),
        )
        if not case.id:
            case.id = TestCase().id
        cases.append(case)
        if no_tc:
            by_no[no_tc] = case

    if "Steps" in book.sheetnames:
        grouped: dict[str, list[dict[str, str]]] = {}
        for item in _sheet_maps(book["Steps"]):
            grouped.setdefault(item.get("no_tc", ""), []).append(item)
        for no_tc, rows in grouped.items():
            case = by_no.get(no_tc)
            if not case:
                continue
            rows.sort(key=lambda row: int(row.get("urutan") or 0))
            case.steps = []
            for item in rows:
                checked = item.get("checked", "").strip().lower()
                checked_val = True if checked in {"true", "1", "yes", "ya"} else False if checked in {"false", "0", "no", "tidak"} else None
                delay = item.get("delay_ms") or "0"
                case.steps.append(
                    Step.from_dict(
                        {
                            "id": item.get("id") or "",
                            "type": item.get("type") or "click",
                            "label": item.get("label", ""),
                            "selector": item.get("selector", ""),
                            "value": item.get("value", ""),
                            "url": item.get("url", ""),
                            "key": item.get("key", ""),
                            "checked": checked_val,
                            "delay_ms": delay,
                            "tag": item.get("tag", ""),
                            "enabled": item.get("enabled", True),
                        }
                    )
                )

    if "Expectations" in book.sheetnames:
        grouped = {}
        for item in _sheet_maps(book["Expectations"]):
            grouped.setdefault(item.get("no_tc", ""), []).append(item)
        for no_tc, rows in grouped.items():
            case = by_no.get(no_tc)
            if not case:
                continue
            rows.sort(key=lambda row: int(row.get("urutan") or 0))
            case.expectations = []
            for item in rows:
                case.expectations.append(
                    Expectation.from_dict(
                        {
                            "id": item.get("id") or "",
                            "label": item.get("label", ""),
                            "selector": item.get("selector", ""),
                            "kind": item.get("kind") or "text",
                            "match": item.get("match") or "contains",
                            "expected_value": item.get("expected_value", ""),
                            "attribute": item.get("attribute", ""),
                            "after_step": item.get("after_step") or 0,
                            "tag": item.get("tag", ""),
                            "sample_text": item.get("sample_text", ""),
                            "enabled": item.get("enabled", True),
                        }
                    )
                )

    name = Path(path).stem
    return TestSuite(name=name, test_cases=cases)
